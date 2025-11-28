#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test script to verify carousel popup content is saved to Excel description field
測試腳本：驗證輪播彈窗內容是否正確存入 Excel 描述欄位
"""

import logging
import sys
from datetime import datetime
from carousel_scraper import CarouselScraper
from document_generator import DocumentGenerator


# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('test_carousel_to_excel.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def main():
    """
    Test carousel scraping and Excel generation with popup content
    """
    carousel_scraper = None
    
    try:
        logger.info("=" * 60)
        logger.info("測試輪播彈窗內容存入 Excel 描述欄位")
        logger.info("=" * 60)
        
        # Configuration
        chromedriver_path = "chromedriver-win64\\chromedriver.exe"
        download_dir = "generated_documents"
        
        # Initialize carousel scraper with DevTools enabled
        logger.info("初始化 CarouselScraper...")
        carousel_scraper = CarouselScraper(
            chromedriver_path=chromedriver_path,
            download_dir=download_dir,
            logger=logger,
            use_chrome_devtools=True
        )
        
        # Set up driver
        logger.info("設定 WebDriver...")
        carousel_scraper.setup_driver()
        
        # Extract carousel banners with popup content
        logger.info("提取輪播橫幅和彈窗內容...")
        banners = carousel_scraper.extract_carousel_banners()
        
        if not banners:
            logger.warning("未找到輪播橫幅")
            return 1
        
        logger.info(f"\n成功提取 {len(banners)} 個輪播橫幅")
        logger.info("=" * 60)
        
        # Display extracted data
        for i, banner in enumerate(banners, 1):
            logger.info(f"\n輪播橫幅 {i}:")
            logger.info(f"  ID: {banner.get('carousel_id', 'N/A')}")
            logger.info(f"  標題: {banner.get('banner_title', 'N/A')}")
            logger.info(f"  圖片URL: {banner.get('image_url', 'N/A')[:50]}...")
            logger.info(f"  課程名稱: {banner.get('course_name', 'N/A')}")
            logger.info(f"  地點: {banner.get('location', 'N/A')}")
            logger.info(f"  講師: {banner.get('instructor', 'N/A')}")
            logger.info(f"  活動連結: {banner.get('activity_link', 'N/A')}")
            
            # Check if description field has popup content
            description = banner.get('description', '')
            if description:
                logger.info(f"  ✓ 描述欄位有內容 ({len(description)} 字元)")
                logger.info(f"  描述內容預覽: {description[:100]}...")
            else:
                logger.warning(f"  ✗ 描述欄位為空")
        
        logger.info("\n" + "=" * 60)
        
        # Initialize document generator
        logger.info("初始化 DocumentGenerator...")
        document_generator = DocumentGenerator(
            output_dir=download_dir,
            logger=logger
        )
        
        # Prepare data for Excel
        logger.info("\n準備輪播資料...")
        excel_data = []
        for banner in banners:
            excel_row = {
                'ID': banner.get('carousel_id', ''),
                '橫幅標題': banner.get('banner_title', ''),
                '圖片URL': banner.get('image_url', ''),
                '活動連結': banner.get('activity_link', ''),
                '課程名稱': banner.get('course_name', ''),
                '地點': banner.get('location', ''),
                '講師': banner.get('instructor', ''),
                '描述': banner.get('description', ''),
                '提取時間': banner.get('extraction_timestamp', '')
            }
            excel_data.append(excel_row)
        
        # Synchronize to Excel
        logger.info("\n將輪播資料同步到 Excel...")
        filename = f"carousel_monitoring_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        success = document_generator.create_monitoring_excel(
            filename=filename,
            content_type='carousel',
            data=excel_data
        )
        
        if success:
            logger.info("✓ Excel 檔案生成成功")
            
            # List generated files
            files = document_generator.list_generated_files()
            if files:
                latest_file = files[0]
                logger.info(f"✓ 最新生成的檔案: {latest_file}")
                
                # Verify Excel content
                logger.info("\n驗證 Excel 內容...")
                verify_excel_content(latest_file, banners)
            else:
                logger.warning("找不到生成的檔案")
        else:
            logger.error("✗ Excel 檔案生成失敗")
            return 1
        
        logger.info("\n" + "=" * 60)
        logger.info("測試完成")
        logger.info("=" * 60)
        
        return 0
        
    except Exception as e:
        logger.error(f"測試過程中發生錯誤: {e}", exc_info=True)
        return 1
        
    finally:
        if carousel_scraper:
            logger.info("清理資源...")
            carousel_scraper.cleanup()


def verify_excel_content(excel_file, expected_data):
    """
    Verify that Excel file contains the expected description content
    
    Args:
        excel_file: Path to Excel file
        expected_data: Expected data list
    """
    try:
        import openpyxl
        
        logger.info(f"讀取 Excel 檔案: {excel_file}")
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active
        
        # Get headers
        headers = [cell.value for cell in worksheet[1]]
        logger.info(f"Excel 欄位: {headers}")
        
        # Find description column index
        description_col_idx = None
        for idx, header in enumerate(headers, 1):
            if header == '描述':
                description_col_idx = idx
                break
        
        if description_col_idx is None:
            logger.warning("找不到「描述」欄位")
            return
        
        logger.info(f"「描述」欄位位於第 {description_col_idx} 欄")
        
        # Check each row
        logger.info("\n檢查 Excel 中的描述欄位內容:")
        for row_idx in range(2, worksheet.max_row + 1):
            description_cell = worksheet.cell(row=row_idx, column=description_col_idx)
            description_value = description_cell.value or ''
            
            if description_value:
                logger.info(f"  第 {row_idx} 列: ✓ 有描述內容 ({len(str(description_value))} 字元)")
                logger.info(f"    內容預覽: {str(description_value)[:80]}...")
            else:
                logger.warning(f"  第 {row_idx} 列: ✗ 描述欄位為空")
        
        workbook.close()
        
        logger.info("\n✓ Excel 內容驗證完成")
        
    except ImportError:
        logger.warning("openpyxl 未安裝，無法驗證 Excel 內容")
    except Exception as e:
        logger.error(f"驗證 Excel 內容時發生錯誤: {e}")


if __name__ == "__main__":
    sys.exit(main())
